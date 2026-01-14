# Excel Report Generator for Annual Trade Reports
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import os
import tempfile
import re

def generate_monthly_excel_report(trade_data, year, month, branch_id=None, branch_name=None):
    """สร้างรายงานรายเดือน (แยกตามวัน)"""
    import calendar
    
    # ชื่อเดือน
    month_names = ['มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 
                   'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
    month_name = month_names[month - 1]
    
    # จำนวนวันในเดือนนั้น
    num_days = calendar.monthrange(year, month)[1]
    
    # 1. เตรียมข้อมูลกราฟ (รายวัน)
    daily_counts = defaultdict(int)
    
    for item in trade_data:
        doc_date = item.get('document_date', '')
        if doc_date:
            # แปลงวันที่จาก format /Date(1704042000000)/ หรือ string
            try:
                dt = parse_date(doc_date)
                if dt.year == year and dt.month == month:
                    daily_counts[dt.day] += 1
            except Exception as e:
                print(f"Warning: Could not parse date {doc_date} in monthly report: {e}")
                pass

    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"รายงานเดือน {month_name}"
    
    # หัวตาราง
    headers = ['วันที่', 'จำนวนรายการ']
    ws.append(headers)
    
    # ข้อมูลตาราง
    days = range(1, num_days + 1)
    categories = [] # สำหรับกราฟ
    values = []     # สำหรับกราฟ
    
    for day in days:
        count = daily_counts[day]
        date_str = f"{day}/{month}/{year}"
        ws.append([date_str, count])
        
        categories.append(str(day))
        values.append(count)
        
    # จัดรูปแบบตาราง
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # สร้างกราฟเส้น (Line Chart)
    line_chart = LineChart()
    line_chart.title = f"สถิติการเทรดรายวัน เดือน {month_name} {year} - {branch_name or branch_id or ''}"
    line_chart.style = 12
    line_chart.y_axis.title = 'จำนวนรายการ'
    line_chart.x_axis.title = 'วันที่'
    
    data = Reference(ws, min_col=2, min_row=1, max_row=num_days + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_days + 1)
    
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    
    ws.add_chart(line_chart, "D2")
    
    # สร้างกราฟแท่ง (Bar Chart)
    bar_chart = BarChart()
    bar_chart.title = "เปรียบเทียบรายวัน"
    bar_chart.style = 10
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    
    ws.add_chart(bar_chart, "D20")
    
    # Save file
    filename = f"monthly_report_{year}_{month:02d}_{branch_id or 'all'}.xlsx"
    # Force /tmp for Vercel
    temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)
    
    return filepath

def generate_annual_excel_report_for_zone(branches_data, year, zone_name, month=None):
    """
    สร้างไฟล์ Excel รายงานรายปี (หรือรายเดือน) สำหรับ Zone - แสดง 3 แถวต่อสาขา
    - branches_data: list ของ dict {branch_id, branch_name, monthly_counts_all, monthly_counts_agreed}
      (หรือ monthly_counts เดิมจะถูกใช้เป็น monthly_counts_all และ agreed = 0)
    - year: ปีที่ต้องการ
    - zone_name: ชื่อ Zone
    - month: เดือนที่ต้องการ (optional)
    """
    wb = Workbook()
    ws = wb.active
    
    month_names_th = ['มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน', 
                   'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
    month_names_short = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.', 
                         'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
    
    if month:
        # รายงานรายเดือน
        import calendar
        month_name = month_names_th[month - 1]
        ws.title = f"Zone {zone_name} - {month_name}"
        
        num_days = calendar.monthrange(year, month)[1]
        headers = ['สาขา', 'รายการ'] + [str(d) for d in range(1, num_days + 1)] + ['รวม']
        data_range = range(1, num_days + 1)
        total_col_idx = num_days + 3  # สาขา(1) + รายการ(2) + days + รวม
    else:
        # รายงานรายปี
        ws.title = f"Report {year}"
        headers = ['สาขา', 'รายการ'] + month_names_short + ['รวม']
        data_range = range(1, 13)
        total_col_idx = 15  # สาขา(1) + รายการ(2) + 12 months + รวม
    
    # --- ส่วนหัวรายงาน (Rows 1-2) ---
    ws['A1'] = f"รายงานรายปี {year}"
    ws['A1'].font = Font(bold=True, size=16, color="333333")
    ws.merge_cells('A1:O1')
    
    ws['A2'] = f"สาขา/Zone: {zone_name}" 
    ws['A2'].font = Font(size=12, color="666666")
    
    # --- ตารางข้อมูล (Row 4) ---
    header_row = 4
    
    # เขียนหัวตาราง
    # คอลัมน์ A = สาขา
    cell = ws.cell(row=header_row, column=1)
    cell.value = 'สาขา'
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # คอลัมน์ B = รายการ
    cell = ws.cell(row=header_row, column=2)
    cell.value = 'รายการ'
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # เขียนชื่อเดือน/วัน (คอลัมน์ 3 เป็นต้นไป)
    if month:
        for col_idx, day_num in enumerate(range(1, num_days + 1), start=3):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = str(day_num)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
    else:
        for col_idx, m_name in enumerate(month_names_short, start=3):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = m_name
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
    
    # คอลัมน์รวม
    cell = ws.cell(row=header_row, column=total_col_idx)
    cell.value = 'รวม'
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- เขียนข้อมูลแต่ละสาขา (3 แถวต่อสาขา) ---
    row_idx = header_row + 1
    
    # สำหรับคำนวณรวมทั้งหมด
    num_periods = len(data_range)
    grand_total_all = [0] * num_periods
    grand_total_agreed = [0] * num_periods
    
    ROW_CONFIGS = [
        {'label': 'ทั้งหมด', 'key_all': True, 'color': 'FFF2CC'},
        {'label': 'ตกลงเทรด', 'key_agreed': True, 'color': 'd4edda'},
        {'label': 'ไม่ตกลง', 'key_not_agreed': True, 'color': 'f8d7da'}
    ]
    
    for branch_idx, branch in enumerate(branches_data):
        branch_name_display = branch.get('branch_name', f"สาขา {branch.get('branch_id')}")
        if ' : ' in branch_name_display:
            branch_name_display = branch_name_display.split(' : ')[-1]
        
        # รองรับ format เดิม (monthly_counts) และ format ใหม่ (monthly_counts_all, monthly_counts_agreed)
        counts_all = branch.get('monthly_counts_all') or branch.get('monthly_counts', {})
        counts_agreed = branch.get('monthly_counts_agreed', {})
        
        for row_config_idx, config in enumerate(ROW_CONFIGS):
            # คอลัมน์ A = ชื่อสาขา (แสดงเฉพาะแถวแรก)
            cell = ws.cell(row=row_idx, column=1)
            if row_config_idx == 0:
                cell.value = branch_name_display
                cell.font = Font(bold=True, size=10)
            else:
                cell.value = ''
            cell.fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            # คอลัมน์ B = รายการ
            cell = ws.cell(row=row_idx, column=2)
            cell.value = config['label']
            cell.font = Font(size=9)
            cell.fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            row_total = 0
            for period_num in data_range:
                col_idx = period_num + 2  # +2 เพราะ column 1=สาขา, 2=รายการ
                
                if config.get('key_all'):
                    count = counts_all.get(period_num, 0)
                    grand_total_all[period_num - 1] += count
                elif config.get('key_agreed'):
                    count = counts_agreed.get(period_num, 0)
                    grand_total_agreed[period_num - 1] += count
                else:  # key_not_agreed
                    count = counts_all.get(period_num, 0) - counts_agreed.get(period_num, 0)
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = count
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                    top=Side(style='thin'), bottom=Side(style='thin'))
                if count > 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                row_total += count
            
            # คอลัมน์รวม
            cell = ws.cell(row=row_idx, column=total_col_idx)
            cell.value = row_total
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            row_idx += 1
    
    # --- แถวรวมทั้งหมด (3 แถว) ---
    for config_idx, config in enumerate(ROW_CONFIGS):
        # คอลัมน์ A = รวมทั้งหมด
        cell = ws.cell(row=row_idx, column=1)
        if config_idx == 0:
            cell.value = 'รวมทั้งหมด'
            cell.font = Font(bold=True, color="FFFFFF")
        else:
            cell.value = ''
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        # คอลัมน์ B = รายการ
        cell = ws.cell(row=row_idx, column=2)
        cell.value = config['label']
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        row_total = 0
        for period_num in data_range:
            col_idx = period_num + 2
            
            if config.get('key_all'):
                count = grand_total_all[period_num - 1]
            elif config.get('key_agreed'):
                count = grand_total_agreed[period_num - 1]
            else:  # not_agreed
                count = grand_total_all[period_num - 1] - grand_total_agreed[period_num - 1]
            
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = count
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            row_total += count
        
        # คอลัมน์รวม
        cell = ws.cell(row=row_idx, column=total_col_idx)
        cell.value = row_total
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        row_idx += 1
    
    # ปรับความกว้างคอลัมน์
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    for col_idx in range(3, total_col_idx + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 7
    
    # สร้างกราฟแท่ง (Bar Chart) - เฉพาะ "ทั้งหมด" ของแต่ละสาขา
    chart = BarChart()
    if month:
        chart.title = f"สถิติการเทรดรายวัน Zone {zone_name} - {month_names_th[month-1]} {year}"
        chart.x_axis.title = 'วันที่'
    else:
        chart.title = f"สถิติการเทรดรายเดือน Zone {zone_name} - {year}"
        chart.x_axis.title = 'เดือน'
        
    chart.y_axis.title = 'จำนวนรายการ'
    chart.style = 10
    chart.height = 12
    chart.width = 20
    
    # วางกราฟ
    ws.add_chart(chart, f"A{row_idx + 3}")
    
    filename = f"zone_report_{year}_{month if month else 'annual'}_{zone_name}.xlsx"
    temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    
    wb.save(filepath)
    print(f"✅ Excel report saved: {filepath}")
    
    return filepath


def generate_annual_excel_report(trade_data, year, branch_id=None, branch_name=None, month=None):
    """
    สร้างไฟล์ Excel รายงานรายปี (หรือรายเดือน) แสดง 3 แถว: ทั้งหมด, ตกลง, ไม่ตกลง
    - trade_data: list ของข้อมูล trade (แต่ละ item มี 'agreed' flag หรือจะดูจาก status)
    - year: ปีที่ต้องการ (ค.ศ.)
    - branch_id: รหัสสาขา (optional)
    - branch_name: ชื่อสาขา (optional)
    - month: เดือนที่ต้องการ (optional) - ถ้าระบุจะเป็นรายงานรายเดือน
    """
    # ถ้ามี month ให้สร้างรายงานรายเดือน (Daily breakdown)
    if month:
        return generate_monthly_excel_report(trade_data, year, month, branch_id, branch_name)

    # 1. เตรียมข้อมูล 3 ประเภท (รายเดือน)
    monthly_counts_all = defaultdict(int)
    monthly_counts_agreed = defaultdict(int)
    
    import re
    
    # สถานะที่ถือว่า "ตกลงเทรด"
    AGREED_STATUSES = ['ยืนยันราคาแล้ว', 'สิ้นสุดการประเมินราคา']
    
    for item in trade_data:
        doc_date = item.get('document_date', '')
        if doc_date:
            try:
                month_num = None
                # รองรับรูปแบบ /Date(timestamp)/
                if doc_date.startswith('/Date('):
                    timestamp_match = re.search(r'/Date\((\d+)\)/', doc_date)
                    if timestamp_match:
                        timestamp = int(timestamp_match.group(1)) / 1000
                        date_obj = datetime.fromtimestamp(timestamp)
                        if date_obj.year == year:
                            month_num = date_obj.month
                else:
                    # รองรับรูปแบบ DD/MM/YYYY
                    date_parts = doc_date.split('/')
                    if len(date_parts) == 3:
                        day, m, year_str = date_parts
                        record_year = int(year_str)
                        if record_year == year:
                            month_num = int(m)
                
                if month_num:
                    # นับทั้งหมด
                    monthly_counts_all[month_num] += 1
                    
                    # ตรวจสอบว่าตกลงเทรดหรือไม่
                    is_agreed = item.get('agreed', False)
                    if not is_agreed:
                        # ลองดูจาก status หรือ BIDDING_STATUS_NAME
                        status = item.get('status')
                        status_name = item.get('BIDDING_STATUS_NAME', '')
                        is_agreed = (status == 3 or status_name in AGREED_STATUSES)
                    
                    if is_agreed:
                        monthly_counts_agreed[month_num] += 1
                        
            except (ValueError, IndexError, AttributeError) as e:
                print(f"Warning: Could not parse date {doc_date}: {e}")
                continue
    
    # สร้าง Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Trade Report {year}"
    
    # กำหนดชื่อเดือน
    month_names = ['ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.', 
                   'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
    
    # --- หัวรายงาน ---
    ws['A1'] = f"รายงานสถิติการเทรด ปี {year}"
    ws['A1'].font = Font(bold=True, size=14, color="333333")
    ws.merge_cells('A1:N1')
    
    if branch_name:
        ws['A2'] = f"สาขา: {branch_name}"
        ws['A2'].font = Font(size=11, color="666666")
    
    # --- หัวตาราง (Row 4) ---
    header_row = 4
    
    # คอลัมน์ A = ประเภท
    cell = ws.cell(row=header_row, column=1)
    cell.value = 'รายการ'
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # เขียนชื่อเดือน (คอลัมน์ 2-13)
    for col_idx, month_name in enumerate(month_names, start=2):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = month_name
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
    
    # คอลัมน์รวม (คอลัมน์ 14)
    cell = ws.cell(row=header_row, column=14)
    cell.value = 'รวม'
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- แถวข้อมูล ---
    row_configs = [
        {'label': 'ทั้งหมด', 'data': monthly_counts_all, 'color': 'FFF2CC'},
        {'label': 'ตกลงเทรด', 'data': monthly_counts_agreed, 'color': 'd4edda'},
        {'label': 'ไม่ตกลง', 'data': None, 'color': 'f8d7da'}  # คำนวณจาก all - agreed
    ]
    
    current_row = header_row + 1
    for config in row_configs:
        # คอลัมน์ A = ชื่อรายการ
        cell = ws.cell(row=current_row, column=1)
        cell.value = config['label']
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        row_total = 0
        for month_num in range(1, 13):
            col_idx = month_num + 1
            
            if config['data'] is not None:
                count = config['data'].get(month_num, 0)
            else:
                # คำนวณ "ไม่ตกลง" = ทั้งหมด - ตกลง
                count = monthly_counts_all.get(month_num, 0) - monthly_counts_agreed.get(month_num, 0)
            
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = count
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            if count > 0:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            
            row_total += count
        
        # คอลัมน์รวม
        cell = ws.cell(row=current_row, column=14)
        cell.value = row_total
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=config['color'], end_color=config['color'], fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        current_row += 1
    
    # ปรับความกว้างคอลัมน์
    ws.column_dimensions['A'].width = 15
    for col_idx in range(2, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 8
    
    # สร้าง Bar Chart (เปรียบเทียบ 3 ประเภท)
    bar_chart = BarChart()
    bar_chart.title = f"สถิติการเทรดรายเดือน - {year}"
    bar_chart.style = 10
    bar_chart.y_axis.title = 'จำนวนรายการ'
    bar_chart.x_axis.title = 'เดือน'
    bar_chart.height = 12
    bar_chart.width = 20
    
    # Data for chart (3 rows of data from row 5-7, columns B-M)
    data = Reference(ws, min_col=2, min_row=header_row, max_col=13, max_row=current_row - 1)
    cats = Reference(ws, min_col=2, min_row=header_row, max_col=13)
    
    bar_chart.add_data(data, titles_from_data=True, from_rows=True)
    bar_chart.set_categories(cats)
    
    # วางกราฟ
    ws.add_chart(bar_chart, "A10")
    
    # สร้างชื่อไฟล์
    if branch_id and branch_name:
        filename = f"trade_report_{year}_branch_{branch_id}.xlsx"
    else:
        filename = f"trade_report_{year}_all_branches.xlsx"
    
    # บันทึกไฟล์ใน temp directory
    temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    
    wb.save(filepath)
    print(f"✅ Excel report saved: {filepath}")
    
    return filepath


def generate_excel_report(trade_data, report_summary, date_start, date_end):
    """
    สร้างรายงาน Excel ตามช่วงวันที่ที่เลือก
    
    Args:
        trade_data: list of trade records
        report_summary: dict of summary data
        date_start: วันที่เริ่มต้น (DD/MM/YYYY)
        date_end: วันที่สิ้นสุด (DD/MM/YYYY)
    
    Returns:
        str: path to generated Excel file
    """
    wb = Workbook()
    
    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # หัวข้อรายงาน
    ws_summary['A1'] = 'Trade Report Summary'
    ws_summary['A1'].font = Font(bold=True, size=16, color="333333")
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f'Date Range: {date_start} - {date_end}'
    ws_summary['A2'].font = Font(size=12, italic=True)
    ws_summary.merge_cells('A2:D2')
    
    # สรุปภาพรวม
    ws_summary['A4'] = 'Overview'
    ws_summary['A4'].font = Font(bold=True, size=14)
    
    overview_data = [
        ('Total Items', report_summary.get('totalCount', 0)),
        ('Confirmed Items', report_summary.get('confirmedCount', 0)),
        ('Not Confirmed Items', report_summary.get('notConfirmedCount', 0)),
        ('Cancelled Items', report_summary.get('cancelledCount', 0)),
        ('Total Amount', f"{report_summary.get('totalAmount', 0):,.2f}"),
        ('Confirmed Amount', f"{report_summary.get('confirmedAmount', 0):,.2f}")
    ]
    
    for idx, (label, value) in enumerate(overview_data, start=5):
        ws_summary.cell(row=idx, column=1).value = label
        ws_summary.cell(row=idx, column=2).value = value
        ws_summary.cell(row=idx, column=1).font = Font(bold=True)
        ws_summary.cell(row=idx, column=2).alignment = Alignment(horizontal='right')
    
    # สรุปตามสถานะ
    row_start = 13
    ws_summary.cell(row=row_start, column=1).value = 'Status Summary'
    ws_summary.cell(row=row_start, column=1).font = Font(bold=True, size=14)
    
    ws_summary.cell(row=row_start+1, column=1).value = 'Status'
    ws_summary.cell(row=row_start+1, column=2).value = 'Count'
    ws_summary.cell(row=row_start+1, column=3).value = 'Amount'
    
    # Header Style
    for col in range(1, 4):
        cell = ws_summary.cell(row=row_start+1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    current_row = row_start + 2
    for status, data in report_summary.get('statusSummary', {}).items():
        ws_summary.cell(row=current_row, column=1).value = status
        ws_summary.cell(row=current_row, column=2).value = data.get('count', 0)
        ws_summary.cell(row=current_row, column=3).value = f"{data.get('amount', 0):,.2f}"
        current_row += 1
        
    # สรุปตามแบรนด์
    row_start = current_row + 2
    ws_summary.cell(row=row_start, column=1).value = 'Brand Summary'
    ws_summary.cell(row=row_start, column=1).font = Font(bold=True, size=14)
    
    ws_summary.cell(row=row_start+1, column=1).value = 'Brand'
    ws_summary.cell(row=row_start+1, column=2).value = 'Count'
    ws_summary.cell(row=row_start+1, column=3).value = 'Amount'
    
    # Header Style
    for col in range(1, 4):
        cell = ws_summary.cell(row=row_start+1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        
    current_row = row_start + 2
    for brand, data in report_summary.get('brandSummary', {}).items():
        ws_summary.cell(row=current_row, column=1).value = brand
        ws_summary.cell(row=current_row, column=2).value = data.get('count', 0)
        ws_summary.cell(row=current_row, column=3).value = f"{data.get('amount', 0):,.2f}"
        current_row += 1
        
    # ปรับความกว้างคอลัมน์ Summary
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 20
    
    # --- Sheet 2: Details ---
    ws_details = wb.create_sheet("Details")
    
    headers = [
        'Document No', 'Date', 'Branch', 'Sale Code', 'Sale Name', 
        'Customer Name', 'Brand', 'Model', 'Status', 'Amount', 
        'Invoice No', 'Ref No'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, item in enumerate(trade_data, start=2):
        ws_details.cell(row=row_idx, column=1).value = item.get('document_no', '')
        ws_details.cell(row=row_idx, column=2).value = item.get('document_date', '')
        ws_details.cell(row=row_idx, column=3).value = item.get('branch_name', '') # Note: API might not return branch_name in item directly if filtered by branch
        ws_details.cell(row=row_idx, column=4).value = item.get('SALE_CODE', '')
        ws_details.cell(row=row_idx, column=5).value = item.get('SALE_NAME', '')
        ws_details.cell(row=row_idx, column=6).value = item.get('customer_name', '')
        ws_details.cell(row=row_idx, column=7).value = item.get('brand_name', '')
        ws_details.cell(row=row_idx, column=8).value = item.get('series', '')
        ws_details.cell(row=row_idx, column=9).value = item.get('BIDDING_STATUS_NAME', '')
        
        amount = item.get('amount')
        try:
            amount = float(amount) if amount else 0.0
        except:
            amount = 0.0
        ws_details.cell(row=row_idx, column=10).value = amount
        
        ws_details.cell(row=row_idx, column=11).value = item.get('invoice_no', '')
        ws_details.cell(row=row_idx, column=12).value = item.get('DOCUMENT_REF_1', '')
    
    # ปรับความกว้างคอลัมน์ Details
    widths = [20, 15, 20, 15, 20, 20, 15, 20, 20, 15, 20, 20]
    for i, w in enumerate(widths):
        ws_details.column_dimensions[chr(65+i)].width = w
        
    # บันทึกไฟล์
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"trade_report_{date_str}.xlsx"
    temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    
    wb.save(filepath)
    print(f"✅ Excel report saved: {filepath}")
    
    return filepath

def parse_year_from_command(year_str):
    """
    แปลงปีจากคำสั่งเป็น Gregorian year
    รองรับทั้ง พ.ศ. และ ค.ศ.
    
    Args:
        year_str: string ปี เช่น "2024", "2567"
    
    Returns:
        int: Gregorian year หรือ None ถ้าไม่ valid
    """
    try:
        year = int(year_str)
        
        # ถ้าเป็น พ.ศ. (มากกว่า 2500) แปลงเป็น ค.ศ.
        if year > 2500:
            year = year - 543
        
        # ตรวจสอบว่าอยู่ในช่วงที่ยอมรับได้
        current_year = datetime.now().year
        if 2020 <= year <= current_year + 1:
            return year
        else:
            return None
    except ValueError:
        return None


def get_year_date_range(year):
    """
    คำนวณวันแรกและวันสุดท้ายของปี
    
    Args:
        year: Gregorian year
    
    Returns:
        tuple: (date_start, date_end) ในรูปแบบ DD/MM/YYYY
    """
    date_start = f"01/01/{year}"
    date_end = f"31/12/{year}"
    
    return date_start, date_end

def parse_date(date_str):
    """
    Helper function to parse date string from various formats
    """
    if not date_str:
        return None
        
    try:
        # Format: /Date(1704042000000)/
        if date_str.startswith('/Date('):
            timestamp_match = re.search(r'/Date\((\d+)\)/', date_str)
            if timestamp_match:
                timestamp = int(timestamp_match.group(1)) / 1000
                return datetime.fromtimestamp(timestamp)
                
        # Format: DD/MM/YYYY
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = map(int, parts)
            return datetime(year, month, day)
            
    except Exception:
        pass
        
    return None
