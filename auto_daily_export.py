# Auto Daily Export Engine
# ระบบ export ข้อมูล trade รายวันอัตโนมัติไป Google Drive

import json
import os
import time
import tempfile
from datetime import datetime, timedelta
from collections import defaultdict

from google_drive_uploader import GoogleDriveUploader


def get_auto_export_config():
    """ดึง config auto-export จาก DB"""
    try:
        from app import get_db_connection
        conn = get_db_connection()
        if not conn:
            return None
        cur = conn.cursor()
        cur.execute("SELECT * FROM auto_export_config ORDER BY id LIMIT 1")
        row = cur.fetchone()
        cur.close()
        conn.close()
        
        if row:
            return dict(row)
        return None
    except Exception as e:
        print(f"❌ Error getting auto-export config: {e}")
        return None


def save_auto_export_log(log_data):
    """บันทึก log การ export"""
    try:
        from app import get_db_connection
        conn = get_db_connection()
        if not conn:
            return
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO auto_export_log 
            (zone_id, zone_name, date_exported, total_records, file_name, 
             gdrive_file_id, status, error_message, duration_seconds)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            log_data.get('zone_id', ''),
            log_data.get('zone_name', ''),
            log_data.get('date_exported', ''),
            log_data.get('total_records', 0),
            log_data.get('file_name', ''),
            log_data.get('gdrive_file_id', ''),
            log_data.get('status', 'success'),
            log_data.get('error_message', ''),
            log_data.get('duration_seconds', 0)
        ))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"❌ Error saving export log: {e}")


def fetch_zone_daily_data(zone, target_date):
    """ดึงข้อมูล trade ของ zone ทั้งหมดในวันที่กำหนด
    Args:
        zone: dict {'zone_id', 'zone_name', 'branch_ids'}
        target_date: datetime date object
    Returns:
        list: trade data items
    """
    from app import fetch_all_for_branch, get_eve_session
    
    date_str = target_date.strftime("%d/%m/%Y")
    branch_ids = zone.get('branch_ids', [])
    all_items = []
    
    print(f"📊 Fetching data for Zone '{zone['zone_name']}' on {date_str}")
    print(f"   Branches: {branch_ids}")
    
    for i, branch_id in enumerate(branch_ids):
        try:
            print(f"   [{i+1}/{len(branch_ids)}] Branch {branch_id}...")
            filters = {
                'date_start': date_str,
                'date_end': date_str,
                'sale_code': '',
                'customer_sign': '',
                'branch_id': branch_id
            }
            items = fetch_all_for_branch(filters)
            
            # เพิ่ม branch_id เข้าไปในทุก item
            for item in items:
                item['_branch_id'] = branch_id
            
            all_items.extend(items)
            print(f"   ✅ Branch {branch_id}: {len(items)} records")
        except Exception as e:
            print(f"   ❌ Branch {branch_id} error: {e}")
    
    print(f"✅ Total records for Zone '{zone['zone_name']}': {len(all_items)}")
    return all_items


def generate_daily_excel(trade_data, zone_name, target_date, branches_info=None):
    """สร้างไฟล์ Excel รายงานรายวัน
    Args:
        trade_data: list of trade records
        zone_name: ชื่อ zone
        target_date: datetime date
        branches_info: dict mapping branch_id -> branch_name
    Returns:
        str: filepath ของไฟล์ Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import re
    
    wb = Workbook()
    ws = wb.active
    
    date_display = target_date.strftime("%d/%m/%Y")
    # Excel sheet title ห้ามมี / \ * ? [ ]
    sheet_title_date = target_date.strftime("%d-%m-%Y")
    ws.title = f"รายงาน {sheet_title_date}"
    
    # === Data Table ===
    # Start from row 1 (No header/footer as requested)
    header_row = 1
    
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    # === Column Definitions ===
    # Standard columns with Thai labels (matching index.html)
    standard_columns = [
        ('branch_id', 'รหัสสาขา'),
        ('branch_name', 'ชื่อสาขา'),
        ('document_no', 'เลขที่คำสั่งเทรด'),
        ('IS_SIGNED', 'ลายเซ็นลูกค้า'),
        ('SIGN_DATE', 'วันที่ลูกค้าคืนเครื่อง'),
        ('document_date', 'วันที่คำสั่งเทรด'),
        ('series', 'สินค้า'),
        ('category_name', 'ประเภทสินค้า'),
        ('brand_name', 'แบรนด์'),
        ('part_number', 'อีมี่/ซีเรียล'),
        ('amount', 'ราคายืนยัน'),
        ('COUPON_TRADE_IN_CODE', 'คูปองส่วนลดค่าเครื่อง'),
        ('invoice_no', 'เลขที่บิลขาย'),
        ('CAMPAIGN_ON_TOP_NAME', 'โปรโมชั่นส่วนลดแบรนด์'),
        ('COUPON_ON_TOP_BRAND_CODE', 'คูปองส่วนลดแบรนด์'),
        ('COUPON_ON_TOP_BRAND_PRICE', 'มูลค่าส่วนลดแบรนด์'),
        ('COUPON_ON_TOP_COMPANY_CODE', 'คูปองส่วนลดบริษัท'),
        ('COUPON_ON_TOP_COMPANY_PRICE', 'มูลค่าส่วนลดบริษัท'),
        ('net_price', 'ราคาสุทธิ'),
        ('customer_name', 'ชื่อผู้ขาย'),
        ('customer_phone_number', 'เบอร์โทรศัพท์ผู้ขาย'),
        ('customer_email', 'อีเมล์ผู้ขาย'),
        ('buyer_name', 'ผู้รับซื้อ'),
        ('SALE_CODE', 'รหัสพนักงานขาย'),
        ('SALE_NAME', 'ชื่อพนักงานขาย'),
        ('DOCUMENT_REF_1', 'เลขที่เอกสารอ้างอิง'),
        ('BIDDING_STATUS_NAME', 'สถานะ'),
        ('CHANGE_REQUEST_COUNT', 'จำนวนที่ถูกแก้ไข'),
        ('trade_in_id', 'Trade In ID')
    ]
    
    # ดึง keys ทั้งหมดจากข้อมูลรายการแรก (ถ้ามี) เพื่อหาคอลัมน์อื่นๆ
    all_api_keys = []
    if trade_data:
        all_api_keys = list(trade_data[0].keys())
        
    # คอลัมน์ที่จะใช้จริง (ลำดับ + Standard + Others)
    final_headers = ['ลำดับ']
    header_to_key = { 'ลำดับ': None }
    
    # 1. ใส่ Standard Columns
    for key, label in standard_columns:
        final_headers.append(label)
        header_to_key[label] = key
        
    # 2. ใส่คอลัมน์ที่เหลือจาก API
    standard_keys = [c[0] for c in standard_columns]
    for key in all_api_keys:
        if key not in standard_keys and key not in ['_branch_id']: 
            final_headers.append(key)
            header_to_key[key] = key

    for col_idx, header in enumerate(final_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # === Data Rows ===
    AGREED_STATUSES = ['ยืนยันราคาแล้ว', 'สิ้นสุดการประเมินราคา']
    agreed_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    cancelled_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    
    total_amount = 0
    total_net_amount = 0
    agreed_count = 0
    
    for row_idx, item in enumerate(trade_data, start=header_row + 1):
        # Calculate derived fields
        bid_internal = str(item.get('_branch_id', item.get('branch_id', '')))
        
        # Get parsed branch info (code and name)
        info = branches_info.get(bid_internal, {'name': bid_internal, 'code': bid_internal})
        item['branch_id'] = info['code']
        item['branch_name'] = info['name']
        
        # Net Price
        try:
            amt = float(item.get('amount', 0) or 0)
            tub = float(item.get('COUPON_ON_TOP_BRAND_PRICE', 0) or 0)
            tuc = float(item.get('COUPON_ON_TOP_COMPANY_PRICE', 0) or 0)
            item['net_price'] = amt + tub + tuc
        except:
            item['net_price'] = float(item.get('amount', 0) or 0)

        status_name = item.get('BIDDING_STATUS_NAME', '')
        is_agreed = status_name in AGREED_STATUSES or item.get('status') == 3
        if is_agreed:
            agreed_count += 1
            total_amount += float(item.get('amount', 0) or 0)
            total_net_amount += item['net_price']

        # Build Row Data
        for col_idx, header in enumerate(final_headers, start=1):
            key = header_to_key[header]
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if header == 'ลำดับ':
                value = row_idx - header_row
            else:
                value = item.get(key, '')
                
                # Special formatting
                if key in ['document_date', 'SIGN_DATE'] and isinstance(value, str) and value.startswith('/Date('):
                    try:
                        ts = int(re.search(r'\d+', value).group()) / 1000
                        value = datetime.fromtimestamp(ts).strftime('%d/%m/%Y %H:%M')
                    except: pass
                elif key == 'IS_SIGNED':
                    value = 'เซ็นแล้ว' if value == '1' else 'ยังไม่เซ็น'
                elif key in ['amount', 'COUPON_ON_TOP_BRAND_PRICE', 'COUPON_ON_TOP_COMPANY_PRICE', 'net_price']:
                    try: value = float(value) if value else 0
                    except: value = 0
            
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            
            # Color coding
            if is_agreed:
                cell.fill = agreed_fill
            elif 'ยกเลิก' in status_name:
                cell.fill = cancelled_fill
                cell.fill = cancelled_fill
    
    # === Summary Row Removed per User Request ===
    
    # === Column widths ===
    widths = [
        6, 20, 15, 20, 20, 
        15, 15, 25, 20, 20, 
        15, 30, 20, 15, 
        20, 15, 15,
        15, 20, 25, 15, 25, 25,
        20, 20, 20, 15,
        15, 25
    ]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    
    # === Save ===
    date_str = target_date.strftime("%Y-%m-%d")
    filename = f"{date_str}_{zone_name}.xlsx"
    temp_dir = '/tmp' if os.path.exists('/tmp') else tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    
    wb.save(filepath)
    print(f"✅ Daily Excel saved: {filepath} ({len(trade_data)} records)")
    
    return filepath


def run_daily_export(force=False):
    """ฟังก์ชันหลัก: Export ข้อมูลรายวันไป Google Drive
    Args:
        force: True = บังคับรัน (ไม่ตรวจสอบว่ารันไปแล้วหรือยัง)
    Returns:
        dict: ผลการรัน
    """
    import pytz
    
    print("\n" + "=" * 60)
    print(f"📤 Auto Daily Export Started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    start_time = time.time()
    
    # 1. ดึง config
    config = get_auto_export_config()
    if not config:
        print("⚠️ No auto-export config found")
        return {'success': False, 'message': 'ไม่พบ config กรุณาตั้งค่าก่อน'}
    
    if not force and not config.get('enabled'):
        print("⚠️ Auto-export is disabled")
        return {'success': False, 'message': 'ระบบ auto-export ปิดอยู่'}
    
    # 2. ตรวจสอบ Google Drive credentials
    gdrive_credentials = config.get('gdrive_credentials', '')
    root_folder_id = config.get('gdrive_folder_id', '')
    max_files = config.get('max_files_per_zone', 365)
    
    if not gdrive_credentials or not root_folder_id:
        print("❌ Google Drive credentials or folder ID not configured")
        return {'success': False, 'message': 'กรุณาตั้งค่า Google Drive credentials และ Folder ID'}
    
    # 3. Initialize Google Drive uploader
    uploader = GoogleDriveUploader(gdrive_credentials)
    test_result = uploader.test_connection()
    if not test_result['success']:
        print(f"❌ Google Drive connection failed: {test_result['message']}")
        return {'success': False, 'message': f'เชื่อมต่อ Google Drive ไม่สำเร็จ: {test_result["message"]}'}
    
    print(f"✅ Google Drive connected: {test_result.get('email', '')}")
    
    # 4. กำหนดวันที่ที่จะ export (วันก่อนหน้า)
    bkk_tz = pytz.timezone('Asia/Bangkok')
    now_bkk = datetime.now(bkk_tz)
    target_date = (now_bkk - timedelta(days=1)).date()
    target_date_dt = datetime.combine(target_date, datetime.min.time())
    
    date_str_display = target_date.strftime("%d/%m/%Y")
    year_month = target_date.strftime("%Y-%m")
    
    print(f"📅 Export date: {date_str_display}")
    
    # 5. ดึง Zone list
    from app import load_custom_zones_from_file, get_branches_from_db
    
    zone_ids_config = config.get('zone_ids', [])
    if isinstance(zone_ids_config, str):
        try:
            zone_ids_config = json.loads(zone_ids_config)
        except:
            zone_ids_config = []
    
    all_zones = load_custom_zones_from_file()
    
    if zone_ids_config:
        # เฉพาะ zone ที่เลือก
        zones_to_export = [z for z in all_zones if z['zone_id'] in zone_ids_config]
    else:
        # ทุก zone
        zones_to_export = all_zones
    
    if not zones_to_export:
        print("⚠️ No zones to export")
        return {'success': False, 'message': 'ไม่มี zone ที่ต้องการ export'}
    
    print(f"📋 Zones to export: {len(zones_to_export)}")
    
    # 6. สร้าง branch name lookup
    branches = get_branches_from_db()
    branches_info = {}
    for b in branches:
        bid = str(b.get('branch_id', ''))
        bname_full = b.get('branch_name', bid)
        
        b_code = bid
        b_name_only = bname_full
        
        if ' : ' in bname_full:
            parts = bname_full.split(' : ')
            # Example: "00115 : ID115 : Studio 7-Future Park-Rangsit"
            # Extract code from first part, name from last part
            raw_code = parts[0].strip()
            if raw_code.isdigit():
                b_code = str(int(raw_code)) # Remove leading zeros if it's all digits
            else:
                b_code = raw_code
                
            b_name_only = parts[-1].strip() if len(parts) > 1 else bname_full
            
        branches_info[bid] = {
            'code': b_code,
            'name': b_name_only
        }
    
    # 7. Export แต่ละ Zone
    results = []
    total_records = 0
    total_files = 0
    total_errors = 0
    
    for zone_idx, zone in enumerate(zones_to_export):
        zone_start = time.time()
        zone_name = zone['zone_name']
        zone_id = zone['zone_id']
        
        print(f"\n{'─' * 40}")
        print(f"📦 [{zone_idx+1}/{len(zones_to_export)}] Processing Zone: {zone_name}")
        
        try:
            # ดึงข้อมูล
            trade_data = fetch_zone_daily_data(zone, target_date_dt)
            
            # สร้าง Excel
            filepath = generate_daily_excel(trade_data, zone_name, target_date_dt, branches_info)
            
            # หารหัส Folder ของ Zone (บันทึกตรงใน Zone folder เลย)
            zone_folder_id = uploader.ensure_folder_path(root_folder_id, zone_name)
            
            if not zone_folder_id:
                raise Exception(f"Failed to find or create folder for zone '{zone_name}'")
            
            # Upload
            filename = f"{target_date.strftime('%Y-%m-%d')}_{zone_name}.xlsx"
            upload_result = uploader.upload_file(filepath, zone_folder_id, filename)
            
            if not upload_result:
                raise Exception("Upload failed")
            
            # FIFO cleanup
            deleted_count = uploader.fifo_cleanup(root_folder_id, zone_name, max_files)
            
            zone_duration = time.time() - zone_start
            
            # บันทึก log
            save_auto_export_log({
                'zone_id': zone_id,
                'zone_name': zone_name,
                'date_exported': date_str_display,
                'total_records': len(trade_data),
                'file_name': filename,
                'gdrive_file_id': upload_result.get('id', ''),
                'status': 'success',
                'duration_seconds': zone_duration
            })
            
            total_records += len(trade_data)
            total_files += 1
            
            results.append({
                'zone_name': zone_name,
                'records': len(trade_data),
                'status': 'success',
                'deleted_old': deleted_count,
                'duration': f"{zone_duration:.1f}s"
            })
            
            print(f"✅ Zone '{zone_name}': {len(trade_data)} records uploaded ({zone_duration:.1f}s)")
            
            # ลบ temp file
            try:
                os.remove(filepath)
            except:
                pass
            
        except Exception as e:
            zone_duration = time.time() - zone_start
            error_msg = str(e)
            print(f"❌ Zone '{zone_name}' failed: {error_msg}")
            
            save_auto_export_log({
                'zone_id': zone_id,
                'zone_name': zone_name,
                'date_exported': date_str_display,
                'total_records': 0,
                'file_name': '',
                'gdrive_file_id': '',
                'status': 'failed',
                'error_message': error_msg,
                'duration_seconds': zone_duration
            })
            
            total_errors += 1
            results.append({
                'zone_name': zone_name,
                'records': 0,
                'status': 'failed',
                'error': error_msg,
                'duration': f"{zone_duration:.1f}s"
            })
    
    total_duration = time.time() - start_time
    
    # 8. ส่ง Telegram notification (ใช้ bot/chat เดียวกับ auto-cancel)
    try:
        from app import get_auto_cancel_config, send_telegram_notification
        
        cancel_config = get_auto_cancel_config()
        if cancel_config:
            bot_token = cancel_config.get('telegram_bot_token', '')
            chat_id = cancel_config.get('telegram_chat_id', '')
            
            if bot_token and chat_id:
                msg = f"""📤 <b>Auto Daily Export Report</b>
📅 ข้อมูลวันที่: {date_str_display}
⏰ เวลารัน: {now_bkk.strftime('%d/%m/%Y %H:%M')}

📊 <b>สรุป:</b>
🗺️ Zone ทั้งหมด: {len(zones_to_export)}
✅ สำเร็จ: {total_files} zone
❌ ล้มเหลว: {total_errors} zone
📋 รายการทั้งหมด: {total_records:,} records
⏱️ ใช้เวลา: {total_duration:.1f} วินาที

📋 <b>รายละเอียด:</b>"""
                
                for r in results:
                    if r['status'] == 'success':
                        msg += f"\n✅ {r['zone_name']}: {r['records']:,} records ({r['duration']})"
                    else:
                        msg += f"\n❌ {r['zone_name']}: {r.get('error', 'unknown error')}"
                
                send_telegram_notification(bot_token, chat_id, msg)
    except Exception as tg_err:
        print(f"⚠️ Telegram notification error: {tg_err}")
    
    print(f"\n{'=' * 60}")
    print(f"📤 Auto Daily Export Completed in {total_duration:.1f}s")
    print(f"   Files: {total_files}, Records: {total_records}, Errors: {total_errors}")
    print(f"{'=' * 60}\n")
    
    return {
        'success': total_errors == 0,
        'total_zones': len(zones_to_export),
        'total_files': total_files,
        'total_records': total_records,
        'total_errors': total_errors,
        'duration': f"{total_duration:.1f}s",
        'results': results
    }
